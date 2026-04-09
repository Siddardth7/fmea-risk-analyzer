"""
test_exporter.py
Tests for src/exporter.py — Excel and PDF export functions.
"""

import io

import pandas as pd
import pytest
import openpyxl

from src.exporter import export_excel


def _sample_df() -> pd.DataFrame:
    """Minimal analyzed FMEA DataFrame (output shape of run_pipeline)."""
    return pd.DataFrame({
        "ID":                     [1, 2, 3],
        "Process_Step":           ["Layup", "Cure", "Demold"],
        "Component":              ["Ply", "Bag", "Part"],
        "Function":               ["F1", "F2", "F3"],
        "Failure_Mode":           ["FM-1", "FM-2", "FM-3"],
        "Effect":                 ["E1", "E2", "E3"],
        "Severity":               [9, 6, 3],
        "Cause":                  ["C1", "C2", "C3"],
        "Occurrence":             [3, 4, 2],
        "Current_Control":        ["Ctrl1", "Ctrl2", "Ctrl3"],
        "Detection":              [4, 5, 2],
        "RPN":                    [108, 120, 12],
        "Risk_Tier":              ["Red", "Red", "Green"],
        "Flag_High_RPN":          [True, True, False],
        "Flag_High_Severity":     [True, False, False],
        "Flag_Action_Priority_H": [True, False, False],
    })


class TestExportExcel:

    def test_returns_bytes(self):
        result = export_excel(_sample_df())
        assert isinstance(result, bytes)

    def test_bytes_non_empty(self):
        result = export_excel(_sample_df())
        assert len(result) > 0

    def test_valid_xlsx_format(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_has_two_sheets(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 2

    def test_sheet_names(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb.sheetnames[0] == "FMEA Analysis"
        assert wb.sheetnames[1] == "Metadata"

    def test_fmea_sheet_row_count(self):
        """Header + 3 data rows = 4 rows total."""
        df = _sample_df()
        result = export_excel(df)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["FMEA Analysis"]
        assert ws.max_row == len(df) + 1  # +1 for header

    def test_fmea_sheet_has_rpn_column(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["FMEA Analysis"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "RPN" in headers

    def test_metadata_sheet_has_content(self):
        result = export_excel(_sample_df())
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Metadata"]
        assert ws.max_row >= 3
