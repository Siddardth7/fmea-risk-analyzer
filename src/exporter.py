"""
exporter.py
FMEA Risk Prioritization Tool — Export Layer

Functions:
    export_excel(df, project_name)                              → bytes  (.xlsx)
    export_pdf(df, pareto_fig, heatmap_fig, project_name, author_name) → bytes  (.pdf)

Both return raw bytes suitable for st.download_button().

Author: Siddardth | M.S. Aerospace Engineering, UIUC
"""

from __future__ import annotations

import io
import os
import tempfile
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------------------------------------------------------------------------
# Shared constants
# ---------------------------------------------------------------------------

_TOOL_VERSION = "1.0.0"
_AIAG_REF     = "AIAG FMEA-4 (4th Ed.) + AIAG/VDA FMEA Handbook (5th Ed., 2019)"

# ---------------------------------------------------------------------------
# Excel palette
# ---------------------------------------------------------------------------

_XL_NAVY       = "1E3A5F"
_XL_BLUE       = "2563EB"
_XL_BLUE_LIGHT = "EFF6FF"
_XL_RED_BG     = "FEE2E2"
_XL_RED_TEXT   = "991B1B"
_XL_AMB_BG     = "FEF3C7"
_XL_AMB_TEXT   = "92400E"
_XL_GRN_BG     = "D1FAE5"
_XL_GRN_TEXT   = "065F46"
_XL_GREY_BG    = "F1F5F9"
_XL_BORDER     = "CBD5E1"
_XL_WHITE      = "FFFFFF"

_TIER_FILLS = {
    "Red":    (PatternFill("solid", fgColor=_XL_RED_BG),  Font(color=_XL_RED_TEXT, size=10)),
    "Yellow": (PatternFill("solid", fgColor=_XL_AMB_BG),  Font(color=_XL_AMB_TEXT, size=10)),
    "Green":  (PatternFill("solid", fgColor=_XL_GRN_BG),  Font(color=_XL_GRN_TEXT, size=10)),
}

_HEADER_FILL  = PatternFill("solid", fgColor=_XL_NAVY)
_HEADER_FONT  = Font(bold=True, color=_XL_WHITE, size=10)
_SUBHEAD_FILL = PatternFill("solid", fgColor=_XL_GREY_BG)
_SUBHEAD_FONT = Font(bold=True, color="1E293B", size=10)
_BOLD_FONT    = Font(bold=True, size=10)
_NORMAL_FONT  = Font(size=10)
_SMALL_FONT   = Font(size=9, color="64748B")
_TITLE_FONT   = Font(bold=True, size=14, color=_XL_NAVY)

_THIN_BORDER = Border(
    left=Side(style="thin", color=_XL_BORDER),
    right=Side(style="thin", color=_XL_BORDER),
    top=Side(style="thin", color=_XL_BORDER),
    bottom=Side(style="thin", color=_XL_BORDER),
)

_CENTER = Alignment(horizontal="center", vertical="center")
_VCENTER = Alignment(vertical="center")
_WRAP   = Alignment(vertical="center", wrap_text=True)

_EXPORT_COLS = [
    "ID", "Process_Step", "Component", "Failure_Mode",
    "Effect", "Severity", "Occurrence", "Detection",
    "RPN", "Risk_Tier",
    "Flag_High_RPN", "Flag_High_Severity", "Flag_Action_Priority_H",
]

_COL_WIDTHS = {
    "ID": 6, "Process_Step": 22, "Component": 18, "Failure_Mode": 32,
    "Effect": 28, "Severity": 10, "Occurrence": 12, "Detection": 11,
    "RPN": 8, "Risk_Tier": 12,
    "Flag_High_RPN": 14, "Flag_High_Severity": 16, "Flag_Action_Priority_H": 20,
}

# ---------------------------------------------------------------------------
# PDF palette / layout
# ---------------------------------------------------------------------------

_PDF_NAVY      = (30, 58, 95)
_PDF_BLUE      = (37, 99, 235)
_PDF_BLUE_LITE = (147, 197, 253)
_PDF_SLATE     = (100, 116, 139)
_PDF_GREY_BG   = (241, 245, 249)
_PDF_WHITE     = (255, 255, 255)
_PDF_TEXT      = (15, 23, 42)
_PDF_RED_BG    = (254, 226, 226)
_PDF_AMB_BG    = (254, 243, 199)
_PDF_GRN_BG    = (209, 250, 229)

_PDF_TIER_RGB = {
    "Red":    _PDF_RED_BG,
    "Yellow": _PDF_AMB_BG,
    "Green":  _PDF_GRN_BG,
}

_PDF_TABLE_COLS = [
    ("ID",           10),
    ("Process Step", 38),
    ("Failure Mode", 52),
    ("S",             8),
    ("O",             8),
    ("D",             8),
    ("RPN",          12),
    ("Tier",         16),
    ("Flags",        65),
]

_PDF_CRITICAL_COLS = [
    ("ID",           10),
    ("Process Step", 35),
    ("Failure Mode", 48),
    ("Cause",        60),
    ("S",             8),
    ("O",             8),
    ("D",             8),
    ("RPN",          14),
    ("Tier",         16),
]

# ---------------------------------------------------------------------------
# Unicode sanitizer
# ---------------------------------------------------------------------------

_UNICODE_MAP = [
    ("\u2014", "-"), ("\u2013", "-"), ("\u00d7", "x"),
    ("\u2265", ">="), ("\u2264", "<="), ("\u00b1", "+/-"),
    ("\u00b0", " deg"), ("\u2018", "'"), ("\u2019", "'"),
    ("\u201c", '"'), ("\u201d", '"'), ("\u2022", "*"),
    ("\u00e9", "e"), ("\u00e0", "a"), ("\u00f1", "n"),
    ("\u00fc", "u"), ("\u00e4", "a"), ("\u00f6", "o"),
]

def _safe(s: object) -> str:
    text = str(s)
    for char, rep in _UNICODE_MAP:
        text = text.replace(char, rep)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _flag_str(row: pd.Series) -> str:
    parts = []
    if row.get("Flag_High_RPN"):          parts.append("High RPN")
    if row.get("Flag_High_Severity"):     parts.append("Sev>=9")
    if row.get("Flag_Action_Priority_H"): parts.append("AP-H")
    return ", ".join(parts) if parts else "-"


# ===========================================================================
# EXCEL EXPORT — 5-Sheet Engineering Workbook
# ===========================================================================

def export_excel(df: pd.DataFrame, project_name: str = "") -> bytes:
    """
    Export to a 5-sheet engineering workbook:
      1. Summary Dashboard — KPIs + risk distribution
      2. Ranked Failure Modes — full styled table
      3. Critical Items — AP-H items + action plan columns
      4. Methodology — FMEA explanation
      5. Metadata — run info
    """
    wb = openpyxl.Workbook()
    _xl_sheet_summary(wb, df, project_name)
    _xl_sheet_ranked(wb, df)
    _xl_sheet_critical(wb, df)
    _xl_sheet_methodology(wb)
    _xl_sheet_metadata(wb, df, project_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sheet 1: Summary Dashboard
# ---------------------------------------------------------------------------

def _xl_sheet_summary(wb: openpyxl.Workbook, df: pd.DataFrame, project_name: str) -> None:
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.sheet_view.showGridLines = False

    now  = datetime.now().strftime("%B %d, %Y  %H:%M")
    proj = project_name or "FMEA Analysis"

    # Title block
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "FMEA Risk Analysis Report"
    c.font  = Font(bold=True, size=18, color=_XL_WHITE)
    c.fill  = PatternFill("solid", fgColor=_XL_NAVY)
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = proj
    c.font  = Font(bold=True, size=13, color=_XL_WHITE)
    c.fill  = PatternFill("solid", fgColor=_XL_NAVY)
    c.alignment = _CENTER
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:G3")
    c = ws["A3"]
    c.value = f"Generated: {now}   |   {_AIAG_REF}"
    c.font  = Font(size=9, color="94A3B8")
    c.fill  = PatternFill("solid", fgColor="1E3A5F")
    c.alignment = _CENTER
    ws.row_dimensions[3].height = 18

    ws.row_dimensions[4].height = 10

    # KPI row labels
    kpis = [
        ("Total Modes",   len(df),                                              _XL_NAVY),
        ("Critical (Red)", int((df["Risk_Tier"] == "Red").sum()),               "991B1B"),
        ("Warning (Yellow)", int((df["Risk_Tier"] == "Yellow").sum()),          "92400E"),
        ("Acceptable (Green)", int((df["Risk_Tier"] == "Green").sum()),         "065F46"),
        ("High RPN (>100)", int(df["Flag_High_RPN"].sum()) if "Flag_High_RPN" in df.columns else 0, "6D28D9"),
        ("Severity >= 9", int(df["Flag_High_Severity"].sum()) if "Flag_High_Severity" in df.columns else 0, "991B1B"),
        ("Action Priority H", int(df["Flag_Action_Priority_H"].sum()) if "Flag_Action_Priority_H" in df.columns else 0, "92400E"),
    ]
    kpi_fills = ["DBEAFE", _XL_RED_BG, _XL_AMB_BG, _XL_GRN_BG, "EDE9FE", _XL_RED_BG, _XL_AMB_BG]

    for i, (label, value, txt_color) in enumerate(kpis, start=1):
        col = get_column_letter(i)
        # label row
        c5 = ws[f"{col}5"]
        c5.value = label
        c5.font  = Font(bold=True, size=9, color=txt_color)
        c5.fill  = PatternFill("solid", fgColor=kpi_fills[i-1])
        c5.alignment = _CENTER
        c5.border = _THIN_BORDER
        ws.row_dimensions[5].height = 20

        # value row
        c6 = ws[f"{col}6"]
        c6.value = value
        c6.font  = Font(bold=True, size=22, color=txt_color)
        c6.fill  = PatternFill("solid", fgColor=kpi_fills[i-1])
        c6.alignment = _CENTER
        c6.border = _THIN_BORDER
        ws.row_dimensions[6].height = 38

    ws.row_dimensions[7].height = 10

    # Risk distribution section
    ws.merge_cells("A8:G8")
    c = ws["A8"]
    c.value = "Risk Distribution"
    c.font  = _SUBHEAD_FONT
    c.fill  = _SUBHEAD_FILL
    c.alignment = _CENTER
    c.border = _THIN_BORDER
    ws.row_dimensions[8].height = 20

    total = len(df)
    dist_rows = [
        ("Red — Immediate Action",   int((df["Risk_Tier"] == "Red").sum()),    _XL_RED_BG,  _XL_RED_TEXT),
        ("Yellow — Action Recommended", int((df["Risk_Tier"] == "Yellow").sum()), _XL_AMB_BG, _XL_AMB_TEXT),
        ("Green — Monitor",          int((df["Risk_Tier"] == "Green").sum()),  _XL_GRN_BG,  _XL_GRN_TEXT),
    ]
    for row_i, (label, count, bg, fg) in enumerate(dist_rows, start=9):
        pct = f"{count/total*100:.1f}%" if total else "0%"
        for col_i, (val, align) in enumerate([
            (label, "left"), (count, "center"), (pct, "center"),
            ("", "center"), ("", "center"), ("", "center"), ("", "center"),
        ], start=1):
            c = ws.cell(row_i, col_i, val)
            c.font   = Font(bold=(col_i <= 2), size=10, color=fg)
            c.fill   = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = _THIN_BORDER
        ws.row_dimensions[row_i].height = 20

    ws.row_dimensions[12].height = 10

    # Top 5 failure modes
    ws.merge_cells("A13:G13")
    c = ws["A13"]
    c.value = "Top 5 Failure Modes by RPN"
    c.font  = _SUBHEAD_FONT
    c.fill  = _SUBHEAD_FILL
    c.alignment = _CENTER
    c.border = _THIN_BORDER
    ws.row_dimensions[13].height = 20

    headers = ["Rank", "Failure Mode", "Process Step", "Component", "S", "O+D", "RPN"]
    for col_i, h in enumerate(headers, start=1):
        c = ws.cell(14, col_i, h)
        c.font  = _HEADER_FONT
        c.fill  = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _THIN_BORDER
    ws.row_dimensions[14].height = 20

    top5 = df.head(5)
    for row_i, (_, row) in enumerate(top5.iterrows(), start=15):
        tier = str(row.get("Risk_Tier", "Green"))
        fill, font = _TIER_FILLS.get(tier, _TIER_FILLS["Green"])
        vals = [
            row_i - 14,
            str(row.get("Failure_Mode", ""))[:40],
            str(row.get("Process_Step", "")),
            str(row.get("Component", "")),
            int(row["Severity"]) if "Severity" in row.index else "",
            (f"{int(row['Occurrence'])}×{int(row['Detection'])}"
             if all(c in row.index for c in ["Occurrence", "Detection"]) else ""),
            int(row["RPN"]) if "RPN" in row.index else "",
        ]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row_i, col_i, val)
            c.fill   = fill
            c.font   = Font(size=10, color=font.color, bold=(col_i == 7))
            c.alignment = _CENTER if col_i in (1, 5, 6, 7) else _VCENTER
            c.border = _THIN_BORDER
        ws.row_dimensions[row_i].height = 20

    # Column widths
    col_widths = [8, 40, 24, 20, 6, 8, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 2: Ranked Failure Modes
# ---------------------------------------------------------------------------

def _xl_sheet_ranked(wb: openpyxl.Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Ranked Failure Modes")
    ws.sheet_view.showGridLines = False

    cols = [c for c in _EXPORT_COLS if c in df.columns]

    # Header row
    for col_i, col_name in enumerate(cols, start=1):
        c = ws.cell(1, col_i, col_name.replace("_", " "))
        c.font      = _HEADER_FONT
        c.fill      = _HEADER_FILL
        c.alignment = _CENTER
        c.border    = _THIN_BORDER
    ws.row_dimensions[1].height = 22

    # Data rows
    for row_i, (_, row) in enumerate(df.iterrows(), start=2):
        tier = str(row.get("Risk_Tier", "Green"))
        fill, font = _TIER_FILLS.get(tier, _TIER_FILLS["Green"])
        for col_i, col_name in enumerate(cols, start=1):
            val = row[col_name]
            if hasattr(val, "item"):
                val = val.item()
            c = ws.cell(row_i, col_i, val)
            c.fill      = fill
            c.font      = Font(size=10, color=font.color)
            c.alignment = _CENTER if col_name in ("ID", "Severity", "Occurrence", "Detection", "RPN") else _VCENTER
            c.border    = _THIN_BORDER
        ws.row_dimensions[row_i].height = 18

    # Column widths + auto-filter
    for col_i, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = _COL_WIDTHS.get(col_name, 14)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 3: Critical Items
# ---------------------------------------------------------------------------

def _xl_sheet_critical(wb: openpyxl.Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Critical Items")
    ws.sheet_view.showGridLines = False

    critical = df[df["Flag_Action_Priority_H"] == True].copy() if "Flag_Action_Priority_H" in df.columns else df.head(0)  # noqa: E712

    # Title
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"Action Priority H Items  ({len(critical)} items require immediate attention)"
    c.font  = Font(bold=True, size=12, color=_XL_WHITE)
    c.fill  = PatternFill("solid", fgColor="991B1B")
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 24

    ws.row_dimensions[2].height = 8

    crit_cols = [
        "ID", "Process_Step", "Component", "Failure_Mode",
        "Cause", "Severity", "Occurrence", "Detection", "RPN", "Risk_Tier",
    ]
    action_cols = ["Assigned To", "Target Date", "Corrective Action", "Status"]
    all_headers = [c for c in crit_cols if c in df.columns] + action_cols

    for col_i, h in enumerate(all_headers, start=1):
        c = ws.cell(3, col_i, h.replace("_", " "))
        c.font      = _HEADER_FONT
        c.fill      = _HEADER_FILL
        c.alignment = _CENTER
        c.border    = _THIN_BORDER
    ws.row_dimensions[3].height = 22

    for row_i, (_, row) in enumerate(critical.iterrows(), start=4):
        tier = str(row.get("Risk_Tier", "Red"))
        fill, font = _TIER_FILLS.get(tier, _TIER_FILLS["Red"])
        data_vals = [row[c] for c in crit_cols if c in df.columns]
        action_vals = ["", "", "", "Open"]
        all_vals = data_vals + action_vals
        for col_i, val in enumerate(all_vals, start=1):
            if hasattr(val, "item"):
                val = val.item()
            c = ws.cell(row_i, col_i, val)
            c.fill      = fill if col_i <= len(data_vals) else PatternFill("solid", fgColor=_XL_GREY_BG)
            c.font      = Font(size=10, color=font.color if col_i <= len(data_vals) else "1E293B")
            c.alignment = _VCENTER
            c.border    = _THIN_BORDER
        ws.row_dimensions[row_i].height = 20

    # Column widths
    widths = [6, 22, 18, 32, 32, 8, 8, 8, 8, 12, 16, 14, 32, 12]
    for i, w in enumerate(widths[:len(all_headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A3:{get_column_letter(len(all_headers))}3"
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet 4: Methodology
# ---------------------------------------------------------------------------

def _xl_sheet_methodology(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False

    content = [
        ("FMEA METHODOLOGY REFERENCE", "title"),
        ("", "blank"),
        ("1. What is FMEA?", "h2"),
        (
            "Failure Mode and Effects Analysis (FMEA) is a systematic, proactive method for evaluating "
            "a process or design to identify where and how it might fail, and to assess the relative "
            "impact of different failures. It enables teams to prioritize corrective actions before "
            "failures occur.",
            "body",
        ),
        ("", "blank"),
        ("2. Risk Priority Number (RPN)", "h2"),
        ("RPN = Severity  x  Occurrence  x  Detection", "formula"),
        ("", "blank"),
        ("Severity (S): 1–10  — How serious is the effect of the failure?", "body"),
        ("  1 = No effect  |  10 = Safety failure without warning (AIAG FMEA-4)", "body"),
        ("Occurrence (O): 1–10  — How often is the cause likely to occur?", "body"),
        ("  1 = Almost never  |  10 = Almost certain (defect rate > 1/2)", "body"),
        ("Detection (D): 1–10  — How well do current controls detect the failure?", "body"),
        ("  1 = Almost certain detection  |  10 = No detection possible", "body"),
        ("", "blank"),
        ("3. Risk Tier Classification", "h2"),
        ("Red   — RPN > 100  OR  Severity >= 9   → Immediate corrective action required", "tier_red"),
        ("Yellow — RPN 50–100  AND  Severity < 9  → Corrective action recommended", "tier_yellow"),
        ("Green  — RPN < 50   AND  Severity < 9   → Monitor; action optional", "tier_green"),
        ("", "blank"),
        ("4. Action Priority (AIAG/VDA 5th Ed.)", "h2"),
        ("Action Priority H (High): RPN >= 200 OR Severity >= 9", "body"),
        ("  Requires immediate escalation, named owner, and documented corrective action plan.", "body"),
        ("", "blank"),
        ("5. Corrective Action Process", "h2"),
        ("1. Assign ownership — named engineer per AP-H item", "body"),
        ("2. Root cause analysis — 5-Why or Ishikawa diagram", "body"),
        ("3. Define actions — reduce Occurrence (process change) or Detection (control upgrade)", "body"),
        ("4. Set deadline — undated action plans are never completed", "body"),
        ("5. Re-score — update S/O/D and verify tier moves to Yellow or Green", "body"),
        ("6. Document — AIAG-compliant traceability of all corrective actions", "body"),
        ("", "blank"),
        ("Engineering Reference", "h2"),
        (_AIAG_REF, "body"),
    ]

    style_map = {
        "title":      (Font(bold=True, size=16, color=_XL_NAVY),  PatternFill("solid", fgColor=_XL_GREY_BG), 32),
        "h2":         (Font(bold=True, size=12, color=_XL_NAVY),  None, 22),
        "formula":    (Font(bold=True, size=12, color="6D28D9"),   PatternFill("solid", fgColor="EDE9FE"), 22),
        "body":       (Font(size=10, color="1E293B"),               None, 18),
        "blank":      (Font(size=6),                                None, 8),
        "tier_red":   (Font(bold=True, size=10, color=_XL_RED_TEXT), PatternFill("solid", fgColor=_XL_RED_BG), 20),
        "tier_yellow":(Font(bold=True, size=10, color=_XL_AMB_TEXT), PatternFill("solid", fgColor=_XL_AMB_BG), 20),
        "tier_green": (Font(bold=True, size=10, color=_XL_GRN_TEXT), PatternFill("solid", fgColor=_XL_GRN_BG), 20),
    }

    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 20

    for row_i, (text, style) in enumerate(content, start=1):
        font, fill, height = style_map[style]
        ws.merge_cells(f"A{row_i}:B{row_i}")
        c = ws[f"A{row_i}"]
        c.value = text
        c.font  = font
        if fill:
            c.fill = fill
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if style == "title" else "left")
        ws.row_dimensions[row_i].height = height


# ---------------------------------------------------------------------------
# Sheet 5: Metadata
# ---------------------------------------------------------------------------

def _xl_sheet_metadata(wb: openpyxl.Workbook, df: pd.DataFrame, project_name: str) -> None:
    ws = wb.create_sheet("Metadata")
    ws.sheet_view.showGridLines = False

    rows = [
        ("Report Title",      "FMEA Risk Analysis Report"),
        ("Project / System",  project_name or "—"),
        ("Generated",         datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Tool Version",      _TOOL_VERSION),
        ("Engineering Ref",   _AIAG_REF),
        ("",                  ""),
        ("Total Rows",        len(df)),
        ("Red (Immediate)",   int((df["Risk_Tier"] == "Red").sum())    if "Risk_Tier"               in df.columns else "N/A"),
        ("Yellow",            int((df["Risk_Tier"] == "Yellow").sum()) if "Risk_Tier"               in df.columns else "N/A"),
        ("Green",             int((df["Risk_Tier"] == "Green").sum())  if "Risk_Tier"               in df.columns else "N/A"),
        ("High RPN (>100)",   int(df["Flag_High_RPN"].sum())           if "Flag_High_RPN"           in df.columns else "N/A"),
        ("Severity >= 9",     int(df["Flag_High_Severity"].sum())      if "Flag_High_Severity"      in df.columns else "N/A"),
        ("Action Priority H", int(df["Flag_Action_Priority_H"].sum())  if "Flag_Action_Priority_H"  in df.columns else "N/A"),
        ("Avg RPN",           f"{df['RPN'].mean():.1f}"                if "RPN"                     in df.columns else "N/A"),
        ("Max RPN",           int(df["RPN"].max())                     if "RPN"                     in df.columns else "N/A"),
    ]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 52

    for r_idx, (label, value) in enumerate(rows, start=1):
        cl = ws.cell(r_idx, 1, label)
        cv = ws.cell(r_idx, 2, value)
        cl.font = _BOLD_FONT if label else _NORMAL_FONT
        cv.font = _NORMAL_FONT
        if label:
            cl.border = _THIN_BORDER
            cv.border = _THIN_BORDER
        ws.row_dimensions[r_idx].height = 18 if label else 8


# ===========================================================================
# PDF EXPORT — 7-Section Industry Report
# ===========================================================================

def export_pdf(
    df: pd.DataFrame,
    pareto_fig: Any,
    heatmap_fig: Any,
    project_name: str = "",
    author_name: str = "",
) -> bytes:
    """
    Export a 7-section professional PDF report:
      1. Cover Page
      2. Executive Summary (KPIs + insights)
      3. Methodology
      4. Ranked Failure Modes (table)
      5. Pareto Chart
      6. Risk Heatmap
      7. Critical Items + Recommendations
    """
    from fpdf import FPDF
    from src.visualizer import pareto_chart as mpl_pareto, risk_heatmap as mpl_heatmap
    import matplotlib.pyplot as plt

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.set_margins(12, 12, 12)

    _pdf_cover(pdf, project_name, author_name)
    _pdf_exec_summary(pdf, df, project_name)
    _pdf_methodology(pdf)
    _pdf_ranked_table(pdf, df)

    for chart_fn, sec_num, title, interp in [
        (
            lambda: mpl_pareto(df),
            "5",
            "Pareto Analysis — Failure Modes Ranked by RPN",
            "Bars are sorted by descending RPN. The cumulative % line identifies the 20% of failure modes "
            "that drive 80% of total risk (Pareto principle). Prioritize corrective resources on failure "
            "modes to the LEFT of the 80% threshold line.",
        ),
        (
            lambda: mpl_heatmap(df),
            "6",
            "Risk Heatmap — Severity x Occurrence Matrix",
            "Each cell shows the count of failure modes at that Severity/Occurrence combination. "
            "Clustering in the top-right corner (high S, high O) indicates systemic process failures "
            "that require immediate process redesign, not just detection improvement.",
        ),
    ]:
        fig = chart_fn()
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            fig.savefig(tmp.name, dpi=150, bbox_inches="tight")
            tmp_path = tmp.name
        plt.close(fig)
        _pdf_chart_page(pdf, tmp_path, sec_num, title, interp)
        os.unlink(tmp_path)

    _pdf_critical_and_recommendations(pdf, df)

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def _pdf_section_header(pdf: Any, num: str, title: str) -> None:
    """Draw a navy section header bar."""
    r, g, b = _PDF_NAVY
    pdf.set_fill_color(r, g, b)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 9, _safe(f"{num}. {title}"), new_x="LMARGIN", new_y="NEXT", fill=True, align="L")
    pdf.ln(3)
    pdf.set_text_color(*_PDF_TEXT)


def _pdf_kpi_row(pdf: Any, df: pd.DataFrame) -> None:
    """Draw a row of KPI boxes."""
    metrics = [
        ("Total Modes",   len(df),                                                                      _PDF_GREY_BG),
        ("Critical (Red)", int((df["Risk_Tier"] == "Red").sum()),                                       (254, 226, 226)),
        ("Warning",       int((df["Risk_Tier"] == "Yellow").sum()),                                     (254, 243, 199)),
        ("Acceptable",    int((df["Risk_Tier"] == "Green").sum()),                                      (209, 250, 229)),
        ("High RPN",      int(df["Flag_High_RPN"].sum()) if "Flag_High_RPN" in df.columns else 0,      _PDF_GREY_BG),
        ("Severity >=9",  int(df["Flag_High_Severity"].sum()) if "Flag_High_Severity" in df.columns else 0, (254, 226, 226)),
        ("Action Prio H", int(df["Flag_Action_Priority_H"].sum()) if "Flag_Action_Priority_H" in df.columns else 0, (254, 243, 199)),
    ]
    cell_w = 272 / len(metrics)

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_text_color(*_PDF_NAVY)
    for label, _, bg in metrics:
        r, g, b = bg
        pdf.set_fill_color(r, g, b)
        pdf.cell(cell_w, 8, _safe(label), border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "B", 14)
    for _, value, bg in metrics:
        r, g, b = bg
        pdf.set_fill_color(r, g, b)
        pdf.cell(cell_w, 11, str(value), border=1, align="C", fill=True)
    pdf.ln(5)
    pdf.set_text_color(*_PDF_TEXT)


def _pdf_cover(pdf: Any, project_name: str, author_name: str) -> None:
    pdf.add_page()

    # ---- Navy header band at top ----
    pdf.set_fill_color(*_PDF_NAVY)
    pdf.rect(0, 0, pdf.w, 42, "F")

    # Blue accent line below header band
    pdf.set_fill_color(*_PDF_BLUE)
    pdf.rect(0, 42, pdf.w, 2.5, "F")

    # Tool label inside header band
    pdf.set_y(12)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(*_PDF_WHITE)
    pdf.cell(0, 7, "FMEA RISK PRIORITIZATION TOOL", align="C")
    pdf.ln(8)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*_PDF_BLUE_LITE)
    pdf.cell(0, 5, _safe(_AIAG_REF), align="C")

    # ---- Main content on white ----
    pdf.set_y(68)

    # Left accent bar beside title
    pdf.set_fill_color(*_PDF_BLUE)
    pdf.rect(pdf.l_margin, pdf.get_y(), 3, 22, "F")

    # Main title
    pdf.set_x(pdf.l_margin + 7)
    pdf.set_font("Helvetica", "B", 26)
    pdf.set_text_color(*_PDF_NAVY)
    pdf.cell(0, 13, "Risk Analysis Report", ln=True)

    pdf.ln(6)

    # Full-width divider
    y_div = pdf.get_y()
    pdf.set_draw_color(*_PDF_BLUE)
    pdf.set_line_width(0.5)
    pdf.line(pdf.l_margin, y_div, pdf.w - pdf.r_margin, y_div)
    pdf.ln(10)

    # Project name
    if project_name:
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(*_PDF_NAVY)
        pdf.cell(0, 9, _safe(project_name), ln=True)
        pdf.ln(4)

    # Date and author
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(*_PDF_SLATE)
    pdf.cell(0, 7, f"Generated: {datetime.now().strftime('%B %d, %Y')}", ln=True)
    if author_name:
        pdf.cell(0, 7, f"Prepared by: {_safe(author_name)}", ln=True)

    # ---- Navy footer band ----
    pdf.set_fill_color(*_PDF_NAVY)
    pdf.rect(0, pdf.h - 16, pdf.w, 16, "F")
    pdf.set_y(pdf.h - 11)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*_PDF_BLUE_LITE)
    pdf.cell(0, 5, "Confidential  |  Engineering Use Only", align="C")


def _pdf_exec_summary(pdf: Any, df: pd.DataFrame, project_name: str) -> None:
    pdf.add_page()
    _pdf_section_header(pdf, "1", "Executive Summary")

    # KPI grid
    _pdf_kpi_row(pdf, df)

    # Auto-insights paragraph
    total_rpn  = df["RPN"].sum()
    top_n      = min(3, len(df))
    top_pct    = df.nlargest(top_n, "RPN")["RPN"].sum() / total_rpn * 100 if total_rpn else 0
    top_item   = df.iloc[0] if not df.empty else None
    red_count  = int((df["Risk_Tier"] == "Red").sum())
    sev9_count = int(df["Flag_High_Severity"].sum()) if "Flag_High_Severity" in df.columns else 0
    ap_h_count = int(df["Flag_Action_Priority_H"].sum()) if "Flag_Action_Priority_H" in df.columns else 0

    insights = []
    if top_item is not None:
        insights.append(
            f"The top {top_n} failure modes account for {top_pct:.0f}% of total RPN."
        )
        insights.append(
            f"Highest-priority failure: '{_safe(str(top_item['Failure_Mode'])[:60])}' "
            f"(RPN = {int(top_item['RPN'])}, {_safe(str(top_item['Process_Step']))})."
        )
    if red_count:
        insights.append(
            f"{red_count} failure mode(s) are in the Red tier and require immediate corrective action."
        )
    if sev9_count:
        insights.append(
            f"{sev9_count} safety-critical failure mode(s) have Severity >= 9 (AIAG FMEA-4 safety rule)."
        )
    if ap_h_count:
        insights.append(
            f"{ap_h_count} item(s) are classified Action Priority H — "
            "each requires a named owner, root cause analysis, and documented action plan."
        )

    r, g, b = (239, 246, 255)
    pdf.set_fill_color(r, g, b)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*_PDF_NAVY)
    pdf.cell(0, 7, "Key Findings", border="B", new_x="LMARGIN", new_y="NEXT",
             fill=True, align="L")
    pdf.ln(1)

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*_PDF_TEXT)
    cw = pdf.w - pdf.l_margin - pdf.r_margin
    for insight in insights:
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(cw, 5, _safe(f"  - {insight}"))
    pdf.ln(4)

    # Top 10 table header
    pdf.set_section_header = True
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*_PDF_NAVY)
    pdf.cell(0, 7, "Top 10 Failure Modes by RPN", new_x="LMARGIN", new_y="NEXT",
             fill=False)
    pdf.ln(1)

    # Widths sum to 186mm = A4(210) - l_margin(12) - r_margin(12)
    top_cols = [("Rank", 8), ("Failure Mode", 64), ("Process Step", 46),
                ("S", 7), ("O", 7), ("D", 7), ("RPN", 11), ("Tier", 16), ("AP-H", 20)]

    pdf.set_fill_color(*_PDF_NAVY)
    pdf.set_text_color(*_PDF_WHITE)
    pdf.set_font("Helvetica", "B", 8)
    for label, w in top_cols:
        pdf.cell(w, 7, label, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for rank_i, (_, row) in enumerate(df.head(10).iterrows(), start=1):
        tier = str(row.get("Risk_Tier", "Green"))
        r2, g2, b2 = _PDF_TIER_RGB.get(tier, _PDF_GREY_BG)
        pdf.set_fill_color(r2, g2, b2)
        pdf.set_text_color(*_PDF_TEXT)
        ap_h = "Yes" if row.get("Flag_Action_Priority_H") else "-"
        vals = [
            (str(rank_i),                                               "C"),
            (_safe(str(row.get("Failure_Mode", ""))[:40]),              "L"),
            (_safe(str(row.get("Process_Step", ""))[:30]),              "L"),
            (str(int(row["Severity"])) if "Severity" in row.index else "","C"),
            (str(int(row["Occurrence"])) if "Occurrence" in row.index else "","C"),
            (str(int(row["Detection"])) if "Detection" in row.index else "","C"),
            (str(int(row["RPN"])) if "RPN" in row.index else "",        "C"),
            (_safe(tier),                                                "C"),
            (ap_h,                                                       "C"),
        ]
        for (val, align), (_, w) in zip(vals, top_cols):
            pdf.cell(w, 6, val, border=1, align=align, fill=True)
        pdf.ln()

    pdf.set_text_color(*_PDF_TEXT)


def _pdf_methodology(pdf: Any) -> None:
    pdf.add_page()
    _pdf_section_header(pdf, "2", "Methodology")

    sections = [
        ("What is FMEA?",
         "Failure Mode and Effects Analysis (FMEA) is a systematic, proactive method for evaluating "
         "a process or design to identify where and how it might fail. It enables engineering teams "
         "to quantify and prioritize risks before failures occur in production."),

        ("Risk Priority Number (RPN)",
         "RPN = Severity x Occurrence x Detection   (range: 1 to 1,000)\n\n"
         "Severity (S):   1-10 — How serious is the effect of the failure?\n"
         "  1 = No effect   |   10 = Safety failure without warning\n\n"
         "Occurrence (O): 1-10 — How often is the cause likely to occur?\n"
         "  1 = Almost never (< 1 in 1,000,000)   |   10 = Almost certain (> 1 in 2)\n\n"
         "Detection (D):  1-10 — How well do current controls detect the failure?\n"
         "  1 = Almost certain detection   |   10 = No detection possible"),

        ("Risk Tier Classification",
         "Red    — RPN > 100  OR  Severity >= 9   →  Immediate corrective action required\n"
         "Yellow — RPN 50-100  AND  Severity < 9   →  Corrective action recommended\n"
         "Green  — RPN < 50   AND  Severity < 9   →  Monitor; action optional"),

        ("Action Priority (AIAG/VDA 5th Ed.)",
         "Action Priority H (High): RPN >= 200 OR Severity >= 9\n"
         "Each AP-H item requires: named owner, root cause analysis (5-Why/Ishikawa),\n"
         "documented corrective action plan with completion date, and re-scoring after action."),

        ("Engineering Reference",
         _AIAG_REF),
    ]

    cw = pdf.w - pdf.l_margin - pdf.r_margin
    for heading, body in sections:
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(*_PDF_NAVY)
        r, g, b = _PDF_GREY_BG
        pdf.set_fill_color(r, g, b)
        pdf.set_x(pdf.l_margin)
        pdf.cell(cw, 8, _safe(heading), new_x="LMARGIN", new_y="NEXT", fill=True, border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*_PDF_TEXT)
        pdf.ln(1)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(cw, 5, _safe(body))
        pdf.ln(5)


def _pdf_ranked_table(pdf: Any, df: pd.DataFrame) -> None:
    pdf.add_page()
    _pdf_section_header(pdf, "3 + 4", "Results — Ranked Failure Modes")

    row_count_label = f"All {len(df)} failure modes ranked by RPN (descending)."
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*_PDF_SLATE)
    pdf.cell(0, 5, _safe(row_count_label), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Table header
    pdf.set_fill_color(*_PDF_NAVY)
    pdf.set_text_color(*_PDF_WHITE)
    pdf.set_font("Helvetica", "B", 8)
    for col_label, col_w in _PDF_TABLE_COLS:
        pdf.cell(col_w, 7, col_label, border=1, align="C", fill=True)
    pdf.ln()

    df2 = df.copy()
    df2["_flags"] = df2.apply(_flag_str, axis=1)
    pdf.set_font("Helvetica", "", 7)

    for _, row in df2.iterrows():
        tier = str(row.get("Risk_Tier", "Green"))
        r, g, b = _PDF_TIER_RGB.get(tier, (255, 255, 255))
        pdf.set_fill_color(r, g, b)
        pdf.set_text_color(*_PDF_TEXT)

        values = [
            str(int(row["ID"]))              if "ID"         in row.index else "",
            _safe(str(row.get("Process_Step", ""))[:22]),
            _safe(str(row.get("Failure_Mode", ""))[:32]),
            str(int(row["Severity"]))        if "Severity"   in row.index else "",
            str(int(row["Occurrence"]))      if "Occurrence" in row.index else "",
            str(int(row["Detection"]))       if "Detection"  in row.index else "",
            str(int(row["RPN"]))             if "RPN"        in row.index else "",
            _safe(tier),
            _safe(str(row.get("_flags", "-"))[:38]),
        ]
        for (_, col_w), val in zip(_PDF_TABLE_COLS, values):
            pdf.cell(col_w, 6, val, border=1,
                     align="C" if col_w <= 16 else "L", fill=True)
        pdf.ln()

    pdf.set_text_color(*_PDF_TEXT)


def _pdf_chart_page(pdf: Any, png_path: str, sec_num: str, title: str, interpretation: str) -> None:
    pdf.add_page()
    _pdf_section_header(pdf, sec_num, title)

    pdf.image(png_path, x=12, w=273)
    pdf.ln(4)

    # Interpretation box
    r, g, b = _PDF_GREY_BG
    pdf.set_fill_color(r, g, b)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_text_color(*_PDF_NAVY)
    pdf.cell(0, 6, "How to read this chart", new_x="LMARGIN", new_y="NEXT",
             fill=True, border="B")
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*_PDF_TEXT)
    pdf.ln(1)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(pdf.w - pdf.l_margin - pdf.r_margin, 4.5, _safe(interpretation))
    pdf.set_text_color(*_PDF_TEXT)


def _pdf_critical_and_recommendations(pdf: Any, df: pd.DataFrame) -> None:
    pdf.add_page()
    _pdf_section_header(pdf, "7", "Critical Items & Recommendations")

    critical = df[df["Flag_Action_Priority_H"] == True] if "Flag_Action_Priority_H" in df.columns else df.head(0)  # noqa: E712

    if critical.empty:
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*_PDF_TEXT)
        pdf.cell(0, 8, "No Action Priority H items found under current filters.", align="L")
        pdf.ln(8)
    else:
        # Warning banner
        pdf.set_fill_color(254, 226, 226)
        pdf.set_text_color(153, 27, 27)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(
            0, 8,
            _safe(f"  {len(critical)} item(s) require immediate corrective action (RPN >= 200 or Severity >= 9)"),
            new_x="LMARGIN", new_y="NEXT", fill=True, border=1, align="L",
        )
        pdf.ln(3)

        # Critical table
        crit_hdrs = [
            ("ID", 10), ("Failure Mode", 55), ("Process Step", 42),
            ("Cause", 56), ("S", 7), ("O", 7), ("D", 7), ("RPN", 12), ("Tier", 17),
        ]
        pdf.set_fill_color(*_PDF_NAVY)
        pdf.set_text_color(*_PDF_WHITE)
        pdf.set_font("Helvetica", "B", 8)
        for label, w in crit_hdrs:
            pdf.cell(w, 7, label, border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 7)
        for _, row in critical.iterrows():
            tier = str(row.get("Risk_Tier", "Red"))
            r, g, b = _PDF_TIER_RGB.get(tier, (255, 255, 255))
            pdf.set_fill_color(r, g, b)
            pdf.set_text_color(*_PDF_TEXT)
            vals = [
                (str(int(row["ID"])) if "ID" in row.index else "",              "C", 10),
                (_safe(str(row.get("Failure_Mode", ""))[:36]),                   "L", 55),
                (_safe(str(row.get("Process_Step", ""))[:28]),                   "L", 42),
                (_safe(str(row.get("Cause", ""))[:38]),                          "L", 56),
                (str(int(row["Severity"])) if "Severity" in row.index else "",   "C",  7),
                (str(int(row["Occurrence"])) if "Occurrence" in row.index else "","C",  7),
                (str(int(row["Detection"])) if "Detection" in row.index else "", "C",  7),
                (str(int(row["RPN"])) if "RPN" in row.index else "",             "C", 12),
                (_safe(tier),                                                     "C", 17),
            ]
            for val, align, w in vals:
                pdf.cell(w, 6, val, border=1, align=align, fill=True)
            pdf.ln()

        pdf.ln(4)
        pdf.set_text_color(*_PDF_TEXT)

    # Recommendations block
    r, g, b = _PDF_GREY_BG
    pdf.set_fill_color(r, g, b)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(*_PDF_NAVY)
    cw = pdf.w - pdf.l_margin - pdf.r_margin
    pdf.set_x(pdf.l_margin)
    pdf.cell(cw, 8, "AIAG FMEA-4 Corrective Action Process",
             new_x="LMARGIN", new_y="NEXT", fill=True, border=1)
    pdf.ln(2)

    steps = [
        ("1. Assign Ownership",
         "Every Action Priority H item must have a named engineer as responsible owner. "
         "Undocumented ownership means corrective action will not happen."),
        ("2. Root Cause Analysis",
         "Conduct a 5-Why or Ishikawa (fishbone) analysis to identify the TRUE root cause. "
         "Do not address symptoms - address the failure mechanism."),
        ("3. Define Corrective Actions",
         "Target Occurrence reduction (process change, mistake-proofing, design modification) "
         "or Detection improvement (enhanced inspection, in-process testing, automated detection)."),
        ("4. Set a Deadline",
         "Action plans without specific completion dates are never completed. "
         "Assign a deadline for each item during the FMEA review meeting."),
        ("5. Re-Score After Action",
         "After implementing the corrective action, update the S/O/D scores and verify the "
         "Risk_Tier moves from Red to Yellow or Green."),
        ("6. Document for Traceability",
         "AIAG-compliant FMEA requires complete traceability: who, what, when, and "
         "evidence of effectiveness. Retain documentation for customer and audit review."),
    ]

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*_PDF_TEXT)
    for step_title, step_body in steps:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*_PDF_NAVY)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(cw, 5, _safe(step_title))
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*_PDF_TEXT)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(cw, 5, _safe(f"  {step_body}"))
        pdf.ln(2)
